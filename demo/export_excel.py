"""Extract + Transform -> Excel (schema review), across all connected platforms.

Same pull -> normalize path as before, extended to loop every connector in the
registry instead of hard-coding Google Ads. Each connector uses its own date
window (see connectors.py — LinkedIn pulls a fixed historical week since its
campaigns went quiet in Nov 2025) and its own 10-per-account / 100-per-connector
sample. Produced so the owner can eyeball the schema before granting Snowflake
access. When the swap happens later, this file is discarded like the endpoint.
"""
import pathlib
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from common.config import WindsorSettings
from etl.connectors import REGISTRY
from etl.normalize import CLEAN_COLUMNS, normalize_row
from etl.sampling import sample_per_account
from windsor.client import WindsorClient, WindsorError

PER_ACCOUNT = 10
CONNECTOR_CAP = 100
OUT_PATH = pathlib.Path("demo/schema_sample_multi.xlsx")

PREFERRED_ORDER = [
    "platform", "account_id", "campaign_id", "campaign_name", "date", "currency",
    "status", "spend", "impressions", "clicks", "conversions", "revenue", "budget", "loaded_at",
]
TEXT_COLS = {"platform", "account_id", "campaign_id", "campaign_name", "currency", "status", "loaded_at"}
MONEY_COLS = {"spend", "revenue", "budget"}
COUNT_COLS = {"impressions", "clicks", "conversions"}

SCHEMA_NOTES = {
    "platform": ("text", "source platform"),
    "account_id": ("text", "kept as text to preserve dashes / avoid number coercion"),
    "campaign_id": ("text", "part of the RAW merge key"),
    "campaign_name": ("text", ""),
    "date": ("date", "reporting day; part of the merge key"),
    "currency": ("text", "account currency — needed for FX in the curated layer"),
    "status": ("text", "campaign status"),
    "spend": ("number", "money, account currency"),
    "impressions": ("number", ""),
    "clicks": ("number", ""),
    "conversions": ("number", ""),
    "revenue": ("number", "conversion value"),
    "budget": ("number", "campaign budget"),
    "loaded_at": ("text", "UTC ISO timestamp of normalization"),
}

_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_BODY_FONT = Font(name="Arial")


def _cell_value(col: str, value):
    if value is None:
        return None
    if col in TEXT_COLS:
        return str(value)
    if isinstance(value, Decimal):
        return float(value)
    return value


def _write_sheet(ws, cols: list[str], rows: list[dict]) -> None:
    for j, col in enumerate(cols, start=1):
        c = ws.cell(row=1, column=j, value=col)
        c.font, c.fill = _HEADER_FONT, _HEADER_FILL
        c.alignment = Alignment(horizontal="left")

    for i, row in enumerate(rows, start=2):
        for j, col in enumerate(cols, start=1):
            c = ws.cell(row=i, column=j, value=_cell_value(col, row.get(col)))
            c.font = _BODY_FONT
            if col == "date":
                c.number_format = "yyyy-mm-dd"
            elif col in MONEY_COLS:
                c.number_format = "#,##0.00"
            elif col in COUNT_COLS:
                c.number_format = "#,##0"

    ws.freeze_panes = "A2"
    for j, col in enumerate(cols, start=1):
        width = max(len(col), *(len(str(r.get(col) or "")) for r in rows), 8) + 2 if rows else len(col) + 2
        ws.column_dimensions[get_column_letter(j)].width = min(width, 40)


def build_workbook(per_connector: dict, out_path: pathlib.Path, notes: dict | None = None) -> pathlib.Path:
    notes = notes or {}
    cols = list(CLEAN_COLUMNS)
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet; add named ones in registry order

    counts = {}
    for spec in REGISTRY.values():
        sampled = sample_per_account(per_connector.get(spec.key, []), PER_ACCOUNT, CONNECTOR_CAP)
        _write_sheet(wb.create_sheet(spec.label[:31]), cols, sampled)
        counts[spec.key] = len(sampled)

    rd = wb.create_sheet("README")
    meta_rows = [
        ("Source", "Windsor Connectors API"),
        ("Sheets", "one per connector; up to 10 campaigns per account, max 100 rows per connector"),
        ("Level", "campaign (one row per campaign, latest day in the pull window)"),
        ("Generated (UTC)", datetime.now(timezone.utc).isoformat(timespec="seconds")),
        ("Note", "RAW landing shape. Empty metrics are blank (NULL), not 0."),
        ("LinkedIn date range", "2025-10-28 to 2025-11-04 — this account's LinkedIn campaigns "
                                 "have had no activity since Nov 2025, so this sample uses its "
                                 "last active week instead of the current date."),
    ]
    for spec in REGISTRY.values():
        note = notes.get(spec.key, "")
        meta_rows.append((f"Rows — {spec.label}", f"{counts[spec.key]}" + (f"  ({note})" if note else "")))
    for i, (k, v) in enumerate(meta_rows, start=1):
        rd.cell(row=i, column=1, value=k).font = Font(name="Arial", bold=True)
        rd.cell(row=i, column=2, value=v).font = _BODY_FONT

    start = len(meta_rows) + 2
    for j, h in enumerate(("Column", "Type", "Notes"), start=1):
        c = rd.cell(row=start, column=j, value=h)
        c.font, c.fill = _HEADER_FONT, _HEADER_FILL
    for i, col in enumerate(cols, start=start + 1):
        t, note = SCHEMA_NOTES.get(col, ("", ""))
        rd.cell(row=i, column=1, value=col).font = _BODY_FONT
        rd.cell(row=i, column=2, value=t).font = _BODY_FONT
        rd.cell(row=i, column=3, value=note).font = _BODY_FONT
    rd.column_dimensions["A"].width = 22
    rd.column_dimensions["B"].width = 12
    rd.column_dimensions["C"].width = 60

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main() -> None:
    s = WindsorSettings()
    client = WindsorClient(s)
    today = date.today()
    per_connector: dict = {}
    notes: dict = {}

    for spec in REGISTRY.values():
        end_date = spec.fixed_date_to or today
        date_from = end_date - timedelta(days=spec.lookback_days)
        try:
            raw_rows = client.get_data(
                spec.windsor_slug,
                fields=list(spec.field_map.keys()),
                date_from=date_from,
                date_to=end_date,
            )
            per_connector[spec.key] = [normalize_row(r, spec) for r in raw_rows]
            print(f"  {spec.label}: pulled {len(raw_rows)}, normalized "
                  f"{len(per_connector[spec.key])}  ({date_from} to {end_date})")
        except WindsorError as exc:
            per_connector[spec.key] = []
            notes[spec.key] = f"pull failed: {type(exc).__name__}"
            print(f"  {spec.label}: {type(exc).__name__} — sheet will be empty")

    out = build_workbook(per_connector, OUT_PATH, notes)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
